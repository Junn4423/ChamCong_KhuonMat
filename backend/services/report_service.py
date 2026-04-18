from datetime import date as date_cls
from datetime import datetime
from io import BytesIO
import re

from backend.models.database import db, Attendance, User


VALID_REPORT_SORT_FIELDS = {
    'date',
    'name',
    'employee_id',
    'department',
    'check_in_time',
    'check_out_time',
    'status',
}

VALID_REPORT_STATUS_FILTERS = {
    'all',
    'present',
    'late',
    'checked_out',
}

STATUS_LABELS = {
    'all': 'tất cả',
    'present': 'đúng giờ',
    'late': 'trễ',
    'checked_out': 'đã checkout',
}

REPORT_SIGNATURE_ROLES = (
    'Người lập biểu',
    'Trưởng bộ phận',
    'Phòng nhân sự',
    'Kế toán tiền lương',
    'Ban giám đốc',
)


def _parse_date(value, fallback):
    text = str(value or '').strip()
    if not text:
        return fallback
    try:
        return datetime.strptime(text, '%Y-%m-%d').date()
    except ValueError:
        return fallback


def _parse_time(value):
    text = str(value or '').strip()
    if not text:
        return None

    for time_format in ('%H:%M', '%H:%M:%S'):
        try:
            return datetime.strptime(text, time_format).time()
        except ValueError:
            continue
    return None


def _safe_text(value):
    return str(value or '').strip()


def _clean_location_text(value):
    raw = _safe_text(value)
    if not raw:
        return ''

    # Legacy payloads can look like: "Address | lat, lng | ±94m".
    # We only keep the first readable location label.
    head = _safe_text(raw.split('|', 1)[0])
    if head:
        return head

    match = re.match(r'^(.*?)\s*\(([-\d.]+),\s*([-\d.]+)(?:\s*[Â±+-](\d+)m)?\)$', raw, re.IGNORECASE)
    if match:
        return _safe_text(match.group(1))

    return raw


def _attendance_status(attendance):
    if attendance.check_out_time:
        return 'checked_out', 'Đã Checkout'
    if attendance.status == 'late':
        return 'late', 'Trễ'
    return 'present', 'Đúng giờ'


def parse_report_filters(source):
    payload = source if isinstance(source, dict) else {}
    today = date_cls.today()
    legacy_date = _safe_text(payload.get('date'))

    start_date = _parse_date(payload.get('start_date') or legacy_date, today)
    end_date = _parse_date(payload.get('end_date') or legacy_date, start_date)
    if end_date < start_date:
        start_date, end_date = end_date, start_date

    start_time = _parse_time(payload.get('start_time'))
    end_time = _parse_time(payload.get('end_time'))
    if start_time and end_time and end_time < start_time:
        start_time, end_time = end_time, start_time

    status = _safe_text(payload.get('status')).lower() or 'all'
    if status not in VALID_REPORT_STATUS_FILTERS:
        status = 'all'

    sort_by = _safe_text(payload.get('sort_by')).lower() or 'check_in_time'
    if sort_by not in VALID_REPORT_SORT_FIELDS:
        sort_by = 'check_in_time'

    sort_dir = _safe_text(payload.get('sort_dir')).lower() or 'desc'
    if sort_dir not in {'asc', 'desc'}:
        sort_dir = 'desc'

    keyword = _safe_text(payload.get('keyword'))

    return {
        'start_date': start_date,
        'end_date': end_date,
        'start_time': start_time,
        'end_time': end_time,
        'status': status,
        'keyword': keyword,
        'sort_by': sort_by,
        'sort_dir': sort_dir,
    }


def _report_row_matches(row, filters):
    status_filter = filters.get('status')
    if status_filter and status_filter != 'all' and row.get('status_key') != status_filter:
        return False

    keyword = _safe_text(filters.get('keyword')).lower()
    if keyword:
        haystack = ' '.join([
            _safe_text(row.get('name')),
            _safe_text(row.get('employee_id')),
            _safe_text(row.get('department')),
        ]).lower()
        if keyword not in haystack:
            return False

    start_time = filters.get('start_time')
    if start_time and row.get('_check_in_clock') and row['_check_in_clock'] < start_time:
        return False

    end_time = filters.get('end_time')
    if end_time and row.get('_check_in_clock') and row['_check_in_clock'] > end_time:
        return False

    return True


def _sort_rows(rows, filters):
    sort_by = filters.get('sort_by') or 'check_in_time'
    reverse = (filters.get('sort_dir') or 'desc') == 'desc'

    def sort_key(row):
        date_sort = row.get('_date_sort') or date_cls.min
        check_in_sort = row.get('_check_in_sort') or datetime.min
        check_out_sort = row.get('_check_out_sort') or datetime.min
        employee_id = _safe_text(row.get('employee_id')).lower()

        if sort_by == 'date':
            return (date_sort, check_in_sort, employee_id)
        if sort_by == 'name':
            return (_safe_text(row.get('name')).lower(), date_sort, check_in_sort)
        if sort_by == 'employee_id':
            return (employee_id, date_sort, check_in_sort)
        if sort_by == 'department':
            return (_safe_text(row.get('department')).lower(), date_sort, check_in_sort)
        if sort_by == 'check_out_time':
            return (check_out_sort, date_sort, employee_id)
        if sort_by == 'status':
            return (_safe_text(row.get('status')).lower(), date_sort, check_in_sort)
        return (check_in_sort, date_sort, employee_id)

    return sorted(rows, key=sort_key, reverse=reverse)


def _build_row(attendance, user):
    from backend.routes._helpers import serialize_attendance_locations

    location_bundle = serialize_attendance_locations(attendance.id)
    checkin_location_payload = (location_bundle or {}).get('checkin')
    checkout_location_payload = (location_bundle or {}).get('checkout')
    status_key, status_text = _attendance_status(attendance)

    return {
        'attendance_id': attendance.id,
        'date': attendance.date.strftime('%Y-%m-%d') if attendance.date else '',
        'date_display': attendance.date.strftime('%d/%m/%Y') if attendance.date else '',
        'name': user.name,
        'employee_id': user.employee_id,
        'department': user.department,
        'check_in_time': attendance.check_in_time.strftime('%H:%M:%S') if attendance.check_in_time else '',
        'check_out_time': attendance.check_out_time.strftime('%H:%M:%S') if attendance.check_out_time else '',
        'check_in_location': _clean_location_text((checkin_location_payload or {}).get('text', '')),
        'check_out_location': _clean_location_text((checkout_location_payload or {}).get('text', '')),
        'status': status_text,
        'status_key': status_key,
        '_attendance': attendance,
        '_user': user,
        '_date_sort': attendance.date or date_cls.min,
        '_check_in_sort': attendance.check_in_time,
        '_check_out_sort': attendance.check_out_time,
        '_check_in_clock': attendance.check_in_time.time() if attendance.check_in_time else None,
    }


def build_report_rows(filters=None):
    normalized_filters = parse_report_filters(filters)

    records = (
        db.session.query(Attendance, User)
        .join(User, Attendance.user_id == User.id)
        .filter(Attendance.date >= normalized_filters['start_date'])
        .filter(Attendance.date <= normalized_filters['end_date'])
        .all()
    )

    rows = []
    for attendance, user in records:
        row = _build_row(attendance, user)
        if _report_row_matches(row, normalized_filters):
            rows.append(row)

    sorted_rows = _sort_rows(rows, normalized_filters)
    summary = {
        'total_records': len(sorted_rows),
        'unique_employees': len({row.get('employee_id') for row in sorted_rows if row.get('employee_id')}),
    }
    return sorted_rows, normalized_filters, summary


def serialize_report_rows(rows):
    serialized_rows = []
    for row in rows:
        serialized_rows.append({
            'attendance_id': row.get('attendance_id'),
            'date': row.get('date'),
            'date_display': row.get('date_display'),
            'name': row.get('name'),
            'employee_id': row.get('employee_id'),
            'department': row.get('department'),
            'check_in_time': row.get('check_in_time'),
            'check_out_time': row.get('check_out_time'),
            'check_in_location': row.get('check_in_location'),
            'check_out_location': row.get('check_out_location'),
            'status': row.get('status'),
            'status_key': row.get('status_key'),
        })
    return serialized_rows


def build_report_filter_text(filters, summary=None):
    normalized = parse_report_filters(filters)
    parts = [
        f"Ngày: {normalized['start_date'].strftime('%d/%m/%Y')} -> {normalized['end_date'].strftime('%d/%m/%Y')}",
    ]

    if normalized.get('start_time') or normalized.get('end_time'):
        from_time = normalized['start_time'].strftime('%H:%M') if normalized.get('start_time') else '--:--'
        to_time = normalized['end_time'].strftime('%H:%M') if normalized.get('end_time') else '--:--'
        parts.append(f"Giờ vào: {from_time} -> {to_time}")

    if normalized.get('status') and normalized['status'] != 'all':
        parts.append(f"Trạng thái: {STATUS_LABELS.get(normalized['status'], normalized['status'])}")

    if normalized.get('keyword'):
        parts.append(f"Từ khóa: {normalized['keyword']}")

    parts.append(f"Sắp xếp: {normalized['sort_by']} {normalized['sort_dir']}")

    if isinstance(summary, dict):
        parts.append(f"Số bản ghi: {int(summary.get('total_records') or 0)}")

    return ' | '.join(parts)


def build_report_workbook(rows, filters, summary):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError('Thiếu thư viện openpyxl để xuất Excel') from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Báo cáo chấm công'
    sheet.sheet_view.showGridLines = False

    header_fill = PatternFill('solid', fgColor='1F4E78')
    title_fill = PatternFill('solid', fgColor='D9EAF7')
    odd_row_fill = PatternFill('solid', fgColor='F8FBFF')
    present_fill = PatternFill('solid', fgColor='E8F5E9')
    late_fill = PatternFill('solid', fgColor='FFF3E0')
    checkout_fill = PatternFill('solid', fgColor='E3F2FD')
    thin_border = Border(
        left=Side(style='thin', color='D0D7DE'),
        right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'),
        bottom=Side(style='thin', color='D0D7DE'),
    )

    sheet.merge_cells('A1:J1')
    sheet['A1'] = 'BÁO CÁO CHẤM CÔNG'
    sheet['A1'].font = Font(size=15, bold=True, color='0F172A')
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A1'].fill = title_fill

    sheet.merge_cells('A2:J2')
    sheet['A2'] = build_report_filter_text(filters, summary)
    sheet['A2'].font = Font(size=10, italic=True, color='334155')
    sheet['A2'].alignment = Alignment(wrap_text=True, vertical='center')

    sheet.merge_cells('A3:J3')
    sheet['A3'] = f"Tạo lúc: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet['A3'].font = Font(size=10, color='475569')
    sheet['A3'].alignment = Alignment(vertical='center')

    headers = [
        'STT',
        'Ngày',
        'Họ tên',
        'Mã NV',
        'Phòng ban',
        'Giờ vào',
        'Giờ ra',
        'Vị trí checkin',
        'Vị trí checkout',
        'Trạng thái',
    ]

    header_row = 5
    for column_index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=title)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for index, row in enumerate(rows, start=1):
        excel_row = header_row + index
        values = [
            index,
            row.get('date_display'),
            row.get('name'),
            row.get('employee_id'),
            row.get('department'),
            row.get('check_in_time'),
            row.get('check_out_time'),
            row.get('check_in_location'),
            row.get('check_out_location'),
            row.get('status'),
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=excel_row, column=column_index, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical='top',
                horizontal='center' if column_index in {1, 2, 6, 7, 10} else 'left',
                wrap_text=column_index in {3, 5, 8, 9, 10},
            )
            if index % 2 == 1:
                cell.fill = odd_row_fill

        status_fill = None
        status_key = row.get('status_key')
        if status_key == 'present':
            status_fill = present_fill
        elif status_key == 'late':
            status_fill = late_fill
        elif status_key == 'checked_out':
            status_fill = checkout_fill

        if status_fill is not None:
            sheet.cell(row=excel_row, column=10).fill = status_fill

    widths = {
        'A': 8,
        'B': 14,
        'C': 28,
        'D': 16,
        'E': 24,
        'F': 12,
        'G': 12,
        'H': 42,
        'I': 42,
        'J': 16,
    }
    for column_name, width in widths.items():
        sheet.column_dimensions[column_name].width = width

    sheet.freeze_panes = 'A6'
    last_data_row = max(header_row + 1, sheet.max_row)
    sheet.auto_filter.ref = f'A5:J{last_data_row}'
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 34

    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = max(sheet.row_dimensions[row_index].height or 20, 20)

    signature_base_row = sheet.max_row + 3
    now = datetime.now()
    date_line = f"Ngày {now.strftime('%d')} tháng {now.strftime('%m')} năm {now.strftime('%Y')}"

    sheet.merge_cells(start_row=signature_base_row, start_column=7, end_row=signature_base_row, end_column=10)
    date_cell = sheet.cell(row=signature_base_row, column=7, value=date_line)
    date_cell.font = Font(size=10, italic=True, color='334155')
    date_cell.alignment = Alignment(horizontal='center', vertical='center')

    signature_title_row = signature_base_row + 1
    signature_note_row = signature_base_row + 2
    signature_name_row = signature_base_row + 7

    for index, role in enumerate(REPORT_SIGNATURE_ROLES):
        start_column = (index * 2) + 1
        end_column = start_column + 1

        sheet.merge_cells(
            start_row=signature_title_row,
            start_column=start_column,
            end_row=signature_title_row,
            end_column=end_column,
        )
        role_cell = sheet.cell(row=signature_title_row, column=start_column, value=role)
        role_cell.font = Font(size=10, bold=True, color='0F172A')
        role_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        sheet.merge_cells(
            start_row=signature_note_row,
            start_column=start_column,
            end_row=signature_note_row,
            end_column=end_column,
        )
        note_cell = sheet.cell(row=signature_note_row, column=start_column, value='(Ký, ghi rõ họ tên)')
        note_cell.font = Font(size=9, italic=True, color='64748B')
        note_cell.alignment = Alignment(horizontal='center', vertical='center')

        sheet.merge_cells(
            start_row=signature_name_row,
            start_column=start_column,
            end_row=signature_name_row,
            end_column=end_column,
        )
        sign_line_cell = sheet.cell(row=signature_name_row, column=start_column, value='')
        sign_line_cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.row_dimensions[signature_base_row].height = 22
    sheet.row_dimensions[signature_title_row].height = 22
    sheet.row_dimensions[signature_note_row].height = 18
    for row_index in range(signature_note_row + 1, signature_name_row):
        sheet.row_dimensions[row_index].height = 20
    sheet.row_dimensions[signature_name_row].height = 22

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
